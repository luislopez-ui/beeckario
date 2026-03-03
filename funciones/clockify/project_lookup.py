from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from difflib import get_close_matches
from functools import lru_cache
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook


_ID_RE = re.compile(r"^[0-9a-fA-F]{24}$")


def _strip_wrapping_quotes(s: str) -> str:
    s = (s or "").strip()
    # Strip common wrapping quotes repeatedly
    quotes = ['"', "'", "“", "”", "‘", "’", "«", "»"]
    changed = True
    while changed and s:
        changed = False
        for q in quotes:
            if s.startswith(q) and s.endswith(q) and len(s) >= 2:
                s = s[1:-1].strip()
                changed = True
    return s


def _normalize_name(s: str) -> str:
    s = _strip_wrapping_quotes(s)
    s = s.lower().strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    # normalize separators a bit (helps token matching)
    s = s.replace("/", " ")
    s = s.replace("|", " ")
    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _tokens(s: str) -> List[str]:
    # Alphanumeric tokens only (robust against punctuation)
    return re.findall(r"[a-z0-9]+", _normalize_name(s))


def _to_bool(val) -> Optional[bool]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    if s in {"si", "sí", "yes", "y", "true", "1"}:
        return True
    if s in {"no", "n", "false", "0"}:
        return False
    return None


def normalize_project_code(raw: str) -> str:
    """Normalize project codes.

    Supports both:
      - 2-part codes:   'NYB 045' / 'NYB-045' / 'nyb045' -> 'NYB.045'
      - 3-part codes:   'AER.MCC.004' / 'AER MCC 004' / 'AER-MCC-004' -> 'AER.MCC.004'

    Notes:
      - Keeps 3 digits for the numeric suffix when length <= 3 (045, 025, 004).
      - Does NOT drop prefixes (fix for 'AER.MCC.004' incorrectly becoming 'MCC.004').
    """
    s = _strip_wrapping_quotes(raw or "").strip().upper()
    if not s:
        return ""

    # Make separators consistent for matching
    s2 = s.replace("/", " ").replace("|", " ")
    s2 = s2.replace("-", " ").replace("_", " ")
    s2 = re.sub(r"\s+", " ", s2).strip()

    # 1) Prefer 3-part pattern: AAA.BBB.123 (separators may be dots or spaces)
    m3 = re.search(r"\b([A-Z]{2,6})\s*\.?\s*([A-Z]{2,6})\s*\.?\s*(\d{2,4})\b", s2)
    if m3:
        p1, p2, num = m3.group(1), m3.group(2), m3.group(3)
        if len(num) <= 3:
            num = num.zfill(3)
        return f"{p1}.{p2}.{num}"

    # 2) 2-part pattern: AAA.123
    m2 = re.search(r"\b([A-Z]{2,6})\s*\.?\s*(\d{2,4})\b", s2)
    if m2:
        prefix, num = m2.group(1), m2.group(2)
        if len(num) <= 3:
            num = num.zfill(3)
        return f"{prefix}.{num}"

    return s



@dataclass(frozen=True)
class ProjectRow:
    """A single row from the Excel project directory."""

    project_name: str  # value from column "Proyecto" (e.g., "NYB.045", "AER.MCC.004")
    project_id: str    # value from column "ID" (Clockify projectId)
    client: Optional[str] = None
    billable_default: Optional[bool] = None
    id_discovery: Optional[str] = None
    id_desarrollo: Optional[str] = None
    id_deployment: Optional[str] = None
    farming: Optional[str] = None
    hunting: Optional[str] = None


@dataclass(frozen=True)
class ProjectMatch:
    project_name: str
    project_id: str
    client: Optional[str] = None
    billable_default: Optional[bool] = None
    id_discovery: Optional[str] = None
    id_desarrollo: Optional[str] = None
    id_deployment: Optional[str] = None
    farming: Optional[str] = None
    hunting: Optional[str] = None
    match_type: str = "exact"  # "id" | "exact" | "contains" | "tokens" | "fuzzy"


def _excel_path() -> Path:
    # .../Beeckario/funciones/clockify/project_lookup.py -> Beeckario root is parents[2]
    root = Path(__file__).resolve().parents[2]
    return root / "directorios" / "clockify_proyectos.xlsx"


def _clean_cell(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


@lru_cache(maxsize=1)
def _load_projects() -> List[ProjectRow]:
    """Load projects from the Excel file in Beeckario/directorios.

    Expected columns (case-insensitive):
      - Proyecto
      - ID
      - Cliente
      - Facturable  (Si/No)
      - ID_Discovery
      - ID_Desarrollo
      - ID_Deployment
      - Farming
      - Hunting
    """
    path = _excel_path()
    if not path.exists():
        raise FileNotFoundError(f"No se encontró el archivo de proyectos: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)

        header = next(rows, None)
        if not header:
            return []

        header_map = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}

        col_proj = header_map.get("proyecto")
        col_id = header_map.get("id") or header_map.get("project id") or header_map.get("projectid")
        col_cliente = header_map.get("cliente")
        col_fact = header_map.get("facturable") or header_map.get("billable") or header_map.get("facturable?")

        if col_proj is None or col_id is None:
            raise ValueError("El archivo clockify_proyectos.xlsx debe contener columnas 'Proyecto' e 'ID'.")

        col_id_disc = header_map.get("id_discovery")
        col_id_dev = header_map.get("id_desarrollo")
        col_id_dep = header_map.get("id_deployment")
        col_farming = header_map.get("farming")
        col_hunting = header_map.get("hunting")

        out: List[ProjectRow] = []
        for r in rows:
            if not r:
                continue
            proj = r[col_proj] if col_proj < len(r) else None
            pid = r[col_id] if col_id < len(r) else None
            if proj is None or pid is None:
                continue

            proj_s = str(proj).strip()
            pid_s = str(pid).strip()
            if not (proj_s and pid_s):
                continue

            cliente_val = r[col_cliente] if (col_cliente is not None and col_cliente < len(r)) else None
            cliente_s = _clean_cell(cliente_val)

            fact_val = r[col_fact] if (col_fact is not None and col_fact < len(r)) else None
            billable_default = _to_bool(fact_val)

            id_disc = r[col_id_disc] if (col_id_disc is not None and col_id_disc < len(r)) else None
            id_dev = r[col_id_dev] if (col_id_dev is not None and col_id_dev < len(r)) else None
            id_dep = r[col_id_dep] if (col_id_dep is not None and col_id_dep < len(r)) else None
            farming = r[col_farming] if (col_farming is not None and col_farming < len(r)) else None
            hunting = r[col_hunting] if (col_hunting is not None and col_hunting < len(r)) else None

            out.append(
                ProjectRow(
                    project_name=proj_s,
                    project_id=pid_s,
                    client=cliente_s,
                    billable_default=billable_default,
                    id_discovery=_clean_cell(id_disc),
                    id_desarrollo=_clean_cell(id_dev),
                    id_deployment=_clean_cell(id_dep),
                    farming=_clean_cell(farming),
                    hunting=_clean_cell(hunting),
                )
            )
        return out
    finally:
        wb.close()


def preload_projects() -> int:
    """Warm the in-memory project cache.

    Returns the number of projects loaded from directorios/clockify_proyectos.xlsx.
    """
    return len(_load_projects())


def looks_like_project_id(value: str) -> bool:
    return bool(_ID_RE.match((value or "").strip()))


def list_clients() -> List[str]:
    """Return distinct client names from the Excel directory."""
    clients = []
    for row in _load_projects():
        cli = row.client
        if cli:
            clients.append(cli)
    # unique preserving order
    seen = set()
    out = []
    for c in clients:
        nc = _normalize_name(c)
        if nc and nc not in seen:
            seen.add(nc)
            out.append(c)
    return out


def projects_by_client(client_query: str) -> List[ProjectMatch]:
    """Return all projects that match a client name (normalized contains match)."""
    q = _normalize_name(client_query or "")
    if not q:
        return []
    out: List[ProjectMatch] = []
    for row in _load_projects():
        cli = row.client
        if cli and q in _normalize_name(cli):
            out.append(
                ProjectMatch(
                    project_name=row.project_name,
                    project_id=row.project_id,
                    client=row.client,
                    billable_default=row.billable_default,
                    id_discovery=row.id_discovery,
                    id_desarrollo=row.id_desarrollo,
                    id_deployment=row.id_deployment,
                    farming=row.farming,
                    hunting=row.hunting,
                    match_type="client",
                )
            )
    return out


def list_projects(client_query: Optional[str] = None, project_query: Optional[str] = None, limit: int = 100) -> List[ProjectMatch]:
    """List projects from the Excel directory (optionally filtered by client/project substring)."""
    cq = _normalize_name(client_query or "")
    pq = _normalize_name(project_query or "")
    out: List[ProjectMatch] = []
    for row in _load_projects():
        if cq and (not row.client or cq not in _normalize_name(row.client)):
            continue
        if pq and pq not in _normalize_name(row.project_name):
            continue
        out.append(
            ProjectMatch(
                project_name=row.project_name,
                project_id=row.project_id,
                client=row.client,
                billable_default=row.billable_default,
                id_discovery=row.id_discovery,
                id_desarrollo=row.id_desarrollo,
                id_deployment=row.id_deployment,
                farming=row.farming,
                hunting=row.hunting,
                match_type="list",
            )
        )
        if len(out) >= limit:
            break
    return out


def find_project_by_id(project_id: str) -> Optional[ProjectMatch]:
    pid = _strip_wrapping_quotes(project_id or "")
    if not pid:
        return None
    for row in _load_projects():
        if (row.project_id or "").strip() == pid.strip():
            return ProjectMatch(
                project_name=row.project_name,
                project_id=row.project_id,
                client=row.client,
                billable_default=row.billable_default,
                id_discovery=row.id_discovery,
                id_desarrollo=row.id_desarrollo,
                id_deployment=row.id_deployment,
                farming=row.farming,
                hunting=row.hunting,
                match_type="id",
            )
    return None


def resolve_project_id(value: str) -> Tuple[Optional[ProjectMatch], List[ProjectMatch]]:
    """Resolve a project identifier from either a 24-hex ID or a name/código.

    Returns:
      (match, candidates)

      - If match is not None -> resolved.
      - If match is None and candidates not empty -> ambiguous (show candidates).
      - If both None/empty -> not found.
    """
    v = _strip_wrapping_quotes(value or "")
    if not v:
        return None, []

    # If it already looks like a Clockify project ID, use it as-is.
    if looks_like_project_id(v):
        return ProjectMatch(project_name=v, project_id=v, match_type="id"), []

    # Normalize common "CODE 045" forms for better matching
    v_norm_code = normalize_project_code(v)
    projects = _load_projects()

    # Build query in normalized text form
    q = _normalize_name(v_norm_code or v)

    # 1) Exact match
    exact: List[ProjectMatch] = []
    for row in projects:
        if _normalize_name(row.project_name) == q:
            exact.append(
                ProjectMatch(
                    project_name=row.project_name,
                    project_id=row.project_id,
                    client=row.client,
                    billable_default=row.billable_default,
                    id_discovery=row.id_discovery,
                    id_desarrollo=row.id_desarrollo,
                    id_deployment=row.id_deployment,
                    farming=row.farming,
                    hunting=row.hunting,
                    match_type="exact",
                )
            )
    if len(exact) == 1:
        return exact[0], []
    if len(exact) > 1:
        return None, sorted(exact, key=lambda m: _normalize_name(m.project_name))[:10]

    # 2) Contains match (substring)
    contains: List[ProjectMatch] = []
    for row in projects:
        nn = _normalize_name(row.project_name)
        if q and q in nn:
            contains.append(
                ProjectMatch(
                    project_name=row.project_name,
                    project_id=row.project_id,
                    client=row.client,
                    billable_default=row.billable_default,
                    id_discovery=row.id_discovery,
                    id_desarrollo=row.id_desarrollo,
                    id_deployment=row.id_deployment,
                    farming=row.farming,
                    hunting=row.hunting,
                    match_type="contains",
                )
            )
    if len(contains) == 1:
        return contains[0], []
    if len(contains) > 1:
        return None, sorted(contains, key=lambda m: _normalize_name(m.project_name))[:10]

    # 3) Token match (query tokens must all be present in project tokens)
    tq = set(_tokens(q))
    if tq:
        token_hits: List[ProjectMatch] = []
        for row in projects:
            tn = set(_tokens(row.project_name))
            if tq.issubset(tn):
                token_hits.append(
                    ProjectMatch(
                        project_name=row.project_name,
                        project_id=row.project_id,
                        client=row.client,
                        billable_default=row.billable_default,
                        id_discovery=row.id_discovery,
                        id_desarrollo=row.id_desarrollo,
                        id_deployment=row.id_deployment,
                        farming=row.farming,
                        hunting=row.hunting,
                        match_type="tokens",
                    )
                )
        if len(token_hits) == 1:
            return token_hits[0], []
        if len(token_hits) > 1:
            return None, sorted(token_hits, key=lambda m: _normalize_name(m.project_name))[:10]

    # 4) Fuzzy match (difflib)
    # Guard: if the user provided a 3-part code (e.g. AER.MCC.004),
    # never "drop" the first prefix via fuzzy matching.
    projects_for_fuzzy = projects
    if (v_norm_code or "").count(".") == 2:
        head = (v_norm_code or "").split(".", 1)[0].upper().strip()
        if head:
            projects_for_fuzzy = [row for row in projects if normalize_project_code(row.project_name).startswith(head + ".")]
            # If we have no candidates with the same head, skip fuzzy completely
            # to avoid resolving AER.MCC.004 -> MCC.004 by accident.
            if not projects_for_fuzzy:
                return None, []

    names_norm = [_normalize_name(row.project_name) for row in projects_for_fuzzy]
    matches = set(get_close_matches(q, names_norm, n=10, cutoff=0.62))
    if matches:
        fuzzy: List[ProjectMatch] = []
        for row in projects_for_fuzzy:
            if _normalize_name(row.project_name) in matches:
                fuzzy.append(
                    ProjectMatch(
                        project_name=row.project_name,
                        project_id=row.project_id,
                        client=row.client,
                        billable_default=row.billable_default,
                        id_discovery=row.id_discovery,
                        id_desarrollo=row.id_desarrollo,
                        id_deployment=row.id_deployment,
                        farming=row.farming,
                        hunting=row.hunting,
                        match_type="fuzzy",
                    )
                )
        if len(fuzzy) == 1:
            return fuzzy[0], []
        if len(fuzzy) > 1:
            return None, sorted(fuzzy, key=lambda m: _normalize_name(m.project_name))[:10]

    return None, []
